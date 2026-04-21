from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
import argparse
import json
import re

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SRC = Path.home() / 'Downloads' / '疫苗及打针记录.xlsx'
DEFAULT_OUT = ROOT / 'data' / 'vaccine_data.json'

MODE_CONFIG = {
    '时间线版本(健康院)': {
        'id': 'health-center',
        'name': '健康院省钱版',
        'description': '尽量在 HK 健康院完成免费疫苗，必要时补充自费项目。',
        'title': '时间线版本（健康院少花钱版）',
        'time_col': None,
        'channels': [
            ('B', 'hk_free', '香港免费-健康院'),
            ('C', 'cn_free', '内地免费-社康'),
            ('D', 'cn_self_domestic', '内地自费-社康（国产）'),
            ('E', 'cn_self_imported', '内地自费-社康（进口）'),
            ('F', 'hk_self', '香港自费（补充）'),
        ],
        'total_cols': [('B', 'hk_free'), ('C', 'cn_free'), ('D', 'cn_self_domestic'), ('E', 'cn_self_imported'), ('F', 'hk_self'), ('G', 'all')],
        'warnings': [
            '原表中“流脑acyw135-nimenrix 2/3”同时出现在 8 月和 12 月，针次文字重复，建议人工复核。',
        ],
    },
    '时间线版本(机构)': {
        'id': 'institution-lite',
        'name': '机构减针版',
        'description': '优先减少针数，在机构自费完成更多组合疫苗。',
        'title': '时间线版本（减针版）',
        'time_col': 'B',
        'channels': [
            ('C', 'hk_free', '香港免费-健康院'),
            ('D', 'hk_self', '香港自费'),
            ('E', 'cn_free', '内地免费-社康'),
            ('F', 'cn_self_domestic', '内地自费-社康（国产）'),
            ('G', 'cn_self_imported', '内地自费-社康（进口）'),
        ],
        'total_cols': [('C', 'hk_free'), ('D', 'hk_self'), ('E', 'cn_free'), ('F', 'cn_self_domestic'), ('G', 'cn_self_imported'), ('H', 'all')],
        'warnings': [
            '原表标题写“共33针”，但针数合计列为 31 针，页面以明细统计和针数合计列为准。',
            '原表中“五合一（百白破+小儿麻痹+HIB）2/3”同时出现在 4 月和 18 月，针次文字重复，建议人工复核。',
        ],
    },
}

AGE_ORDER = {
    '出生': 0,
    '1月': 1,
    '2月': 2,
    '3月': 3,
    '4月': 4,
    '6月': 6,
    '8月': 8,
    '12月': 12,
    '18月': 18,
    '每年日常': 24,
    '2岁': 30,
    '6岁': 40,
    '11岁': 50,
    '12岁': 60,
}

CHANNEL_THEME = {
    'hk_free': 'free',
    'hk_self': 'purple',
    'cn_free': 'green',
    'cn_self_domestic': 'orange',
    'cn_self_imported': 'gold',
}


def clean_text(value):
    if value is None:
        return None
    text = str(value).replace('🟦', '').replace('\n', ' ').strip()
    text = re.sub(r'\s+', ' ', text)
    return text


def to_display(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return clean_text(value)


def slugify(text):
    base = re.sub(r'[^\w\u4e00-\u9fff]+', '-', clean_text(text) or '')
    return base.strip('-').lower() or 'item'


def extract_declared_total(title):
    match = re.search(r'共(\d+)针', title or '')
    return int(match.group(1)) if match else None


def parse_mode(ws, config):
    groups = []
    current_group = None
    current_age = None
    current_window = None
    item_seq = 0

    for row in range(4, ws.max_row + 1):
        age_val = ws[f'A{row}'].value
        if age_val == '针数':
            break
        if age_val is not None:
            current_age = clean_text(age_val)
            current_window = to_display(ws[f"{config['time_col']}{row}"].value) if config['time_col'] else None
            current_group = {
                'id': f"{config['id']}-{slugify(current_age)}",
                'age': current_age,
                'order': AGE_ORDER.get(current_age, 999),
                'window': current_window,
                'items': [],
            }
            groups.append(current_group)
        elif config['time_col'] and current_group and ws[f"{config['time_col']}{row}"].value is not None:
            current_window = to_display(ws[f"{config['time_col']}{row}"].value)
            current_group['window'] = current_window

        if not current_group:
            continue

        for col, channel_key, channel_label in config['channels']:
            raw = ws[f'{col}{row}'].value
            if raw is None:
                continue
            item_seq += 1
            label = clean_text(raw)
            current_group['items'].append({
                'id': f"{config['id']}-{item_seq}",
                'label': label,
                'rawLabel': str(raw),
                'channelKey': channel_key,
                'channel': channel_label,
                'channelTheme': CHANNEL_THEME[channel_key],
                'age': current_group['age'],
                'window': current_group['window'],
            })

    title_cell = clean_text(ws['A1'].value)
    declared_total = extract_declared_total(title_cell)
    totals = {}
    for col, key in config['total_cols']:
        value = ws[f'{col}26'].value
        totals[key] = int(value) if isinstance(value, (int, float)) else None

    summary = {
        'id': config['id'],
        'name': config['name'],
        'title': config['title'],
        'description': config['description'],
        'declaredTotal': declared_total,
        'sheetTotal': totals.get('all'),
        'totalsByChannel': {k: v for k, v in totals.items() if k != 'all'},
        'groupCount': len(groups),
        'itemCount': sum(len(group['items']) for group in groups),
        'warnings': list(config['warnings']),
    }
    if declared_total and totals.get('all') and declared_total != totals.get('all'):
        summary['warnings'].insert(0, f"标题标注 {declared_total} 针，但针数合计列是 {totals.get('all')} 针。")

    return summary, groups


def build_payload(src: Path):
    wb = load_workbook(src, data_only=True)
    summaries = []
    modes = []
    top_warnings = []

    for sheet_name in ['时间线版本(健康院)', '时间线版本(机构)']:
        summary, groups = parse_mode(wb[sheet_name], MODE_CONFIG[sheet_name])
        summaries.append(summary)
        modes.append({
            'id': summary['id'],
            'name': summary['name'],
            'title': summary['title'],
            'description': summary['description'],
            'declaredTotal': summary['declaredTotal'],
            'sheetTotal': summary['sheetTotal'],
            'totalsByChannel': summary['totalsByChannel'],
            'warnings': summary['warnings'],
            'groups': groups,
        })
        for warning in summary['warnings']:
            top_warnings.append({'modeId': summary['id'], 'message': warning})

    birth_date = to_display(wb['时间线版本(机构)']['B4'].value)

    return {
        'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'sourceWorkbook': str(src),
        'baby': {
            'birthDate': birth_date,
        },
        'comparison': summaries,
        'warnings': top_warnings,
        'modes': modes,
    }


def parse_args():
    parser = argparse.ArgumentParser(description='从疫苗 Excel 抽取 GitHub Pages 页面所需 JSON 数据。')
    parser.add_argument('--src', default=str(DEFAULT_SRC), help='原始 Excel 路径，默认使用 ~/Downloads/疫苗及打针记录.xlsx')
    parser.add_argument('--out', default=str(DEFAULT_OUT), help='输出 JSON 路径，默认写入 ../data/vaccine_data.json')
    return parser.parse_args()


def main():
    args = parse_args()
    src = Path(args.src).expanduser().resolve()
    out = Path(args.out).expanduser().resolve()

    if not src.exists():
        raise FileNotFoundError(f'未找到 Excel 文件：{src}。请通过 --src 传入正确路径。')

    out.parent.mkdir(parents=True, exist_ok=True)
    payload = build_payload(src)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    print(out)


if __name__ == '__main__':
    main()
